"""Excel-Export für Analyse-Ergebnisse — Massenermittlung als .xlsx."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
WARN_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FMT_M2 = '#,##0.00" m²"'
NUM_FMT_M = '#,##0.00" m"'
NUM_FMT_PCT = '0"%"'


def generate_excel(result: dict, filename: str = "Massenermittlung") -> bytes:
    """
    Generates an Excel workbook from an analysis result.
    Returns the .xlsx as bytes.
    """
    wb = Workbook()

    # Sheet 1: Übersicht
    _create_overview_sheet(wb, result, filename)

    # Sheet 2: Räume
    if result.get("raeume"):
        _create_raeume_sheet(wb, result["raeume"])

    # Sheet 3: Decken
    if result.get("decken"):
        _create_decken_sheet(wb, result["decken"])

    # Sheet 4: Wände
    if result.get("waende"):
        _create_waende_sheet(wb, result["waende"])

    # Sheet 5: Warnungen
    if result.get("warnungen"):
        _create_warnungen_sheet(wb, result["warnungen"])

    # Remove default empty sheet if others exist
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 40)


def _create_overview_sheet(wb: Workbook, result: dict, filename: str) -> None:
    ws = wb.active
    ws.title = "Übersicht"

    ws["A1"] = "LaneCore AI — Massenermittlung"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    rows = [
        ("Datei:", filename),
        ("Plantyp:", result.get("plantyp", "—")),
        ("Maßstab:", result.get("massstab", "—")),
        ("Geschoss:", result.get("geschoss", "—")),
        ("Konfidenz:", f"{(result.get('konfidenz', 0) or 0) * 100:.0f}%"),
        ("Exportiert:", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("", ""),
        ("Zusammenfassung:", ""),
    ]

    summary = result.get("summary", {})
    rows.append(("Anzahl Räume:", summary.get("anzahl_raeume", len(result.get("raeume", [])))))
    rows.append(("Raumfläche gesamt:", f"{summary.get('gesamt_raumflaeche', 0):.2f} m²"))

    # Wandflächen nach Typ
    wf = summary.get("gesamt_wandflaeche", {})
    if isinstance(wf, dict):
        for typ, flaeche in wf.items():
            if flaeche and flaeche > 0:
                rows.append((f"Wandfläche {typ}:", f"{flaeche:.2f} m²"))

    # Deckenflächen nach System
    df = summary.get("gesamt_deckenflaeche", {})
    if isinstance(df, dict):
        for sys, flaeche in df.items():
            if flaeche and flaeche > 0:
                rows.append((f"Deckenfläche {sys}:", f"{flaeche:.2f} m²"))

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30


def _create_raeume_sheet(wb: Workbook, raeume: list) -> None:
    ws = wb.create_sheet("Räume")
    headers = ["Raum-Nr", "Bezeichnung", "Fläche (m²)", "Höhe (m)", "Nutzung", "Deckentyp"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in raeume:
        ws.append([
            r.get("raum_nr", ""),
            r.get("bezeichnung", ""),
            r.get("flaeche_m2"),
            r.get("hoehe_m"),
            r.get("nutzung", ""),
            r.get("deckentyp", ""),
        ])

    # Format number columns
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value:
                cell.number_format = NUM_FMT_M2
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                cell.number_format = NUM_FMT_M

    _auto_width(ws)


def _create_decken_sheet(wb: Workbook, decken: list) -> None:
    ws = wb.create_sheet("Decken")
    headers = ["Raum", "Deckentyp", "System", "Fläche (m²)", "Abhängehöhe (m)", "Profil", "Entfällt"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for d in decken:
        entfaellt = "JA" if d.get("entfaellt") else ""
        row_num = ws.max_row + 1
        ws.append([
            d.get("raum", ""),
            d.get("typ", ""),
            d.get("system", ""),
            d.get("flaeche_m2"),
            d.get("abhaengehoehe_m"),
            d.get("profil", ""),
            entfaellt,
        ])
        if d.get("entfaellt"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = WARN_FILL

    _auto_width(ws)


def _create_waende_sheet(wb: Workbook, waende: list) -> None:
    ws = wb.create_sheet("Wände")
    headers = ["ID", "Wandtyp", "Länge (m)", "Höhe (m)", "Fläche (m²)", "Von Raum", "Zu Raum", "Notizen"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for w in waende:
        laenge = w.get("laenge_m") or 0
        hoehe = w.get("hoehe_m") or 0
        flaeche = w.get("flaeche_m2") or (laenge * hoehe if laenge and hoehe else None)
        ws.append([
            w.get("id", ""),
            w.get("typ", ""),
            laenge or None,
            hoehe or None,
            flaeche,
            w.get("von_raum", ""),
            w.get("zu_raum", ""),
            w.get("notizen", ""),
        ])

    _auto_width(ws)


def _create_warnungen_sheet(wb: Workbook, warnungen: list) -> None:
    ws = wb.create_sheet("Warnungen")
    ws.append(["Nr.", "Warnung"])
    _style_header_row(ws, 2)

    for i, w in enumerate(warnungen, 1):
        ws.append([i, w])
        ws.cell(row=i + 1, column=2).fill = WARN_FILL

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80
